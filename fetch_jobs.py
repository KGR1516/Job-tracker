"""
Daily Job Openings Tracker
Fetches job postings via the JSearch API (RapidAPI) and saves them
into a dated Excel file inside the /reports folder.
Runs automatically every day via GitHub Actions.
"""

import os
import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------- CONFIG - EDIT THESE ----------------
SEARCH_QUERIES = [
    "SAP FICO Consultant in UAE",
    "Accountant in Dubai",
    "Financial Analyst in Sharjah",
]
PAGES_PER_QUERY = 1          # each page = ~10 jobs
DATE_POSTED = "today"        # options: all, today, 3days, week, month
# -----------------------------------------------------

API_KEY = os.environ.get("RAPIDAPI_KEY")
API_URL = "https://jsearch.p.rapidapi.com/search-v2"


def fetch_jobs(query: str) -> list:
    headers = {
        "X-RapidAPI-Key": API_KEY,
        "X-RapidAPI-Host": "jsearch.p.rapidapi.com",
    }
    jobs = []
    for page in range(1, PAGES_PER_QUERY + 1):
        params = {
            "query": query,
            "page": page,
            "num_pages": 1,
            "date_posted": DATE_POSTED,
        }
        resp = requests.get(API_URL, headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("data", [])
        for j in data:
            jobs.append({
                "Search": query,
                "Job Title": j.get("job_title", ""),
                "Company": j.get("employer_name", ""),
                "Location": ", ".join(filter(None, [j.get("job_city"), j.get("job_country")])),
                "Employment Type": j.get("job_employment_type", ""),
                "Posted": (j.get("job_posted_at_datetime_utc") or "")[:10],
                "Salary": format_salary(j),
                "Source": j.get("job_publisher", ""),
                "Apply Link": j.get("job_apply_link", ""),
            })
    return jobs


def format_salary(j: dict) -> str:
    lo, hi = j.get("job_min_salary"), j.get("job_max_salary")
    cur = j.get("job_salary_currency") or ""
    if lo and hi:
        return f"{cur} {lo:,.0f} - {hi:,.0f}"
    if lo:
        return f"{cur} {lo:,.0f}+"
    return ""


def dedupe(jobs: list) -> list:
    seen, unique = set(), []
    for j in jobs:
        key = (j["Job Title"].lower(), j["Company"].lower())
        if key not in seen:
            seen.add(key)
            unique.append(j)
    return unique


def write_excel(jobs: list, path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Openings"

    headers = ["Search", "Job Title", "Company", "Location", "Employment Type",
               "Posted", "Salary", "Source", "Apply Link"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for j in jobs:
        ws.append([j[h] for h in headers])

    # Make apply links clickable
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=len(headers))
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")

    # Column widths
    widths = [28, 40, 28, 24, 16, 12, 20, 16, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main():
    if not API_KEY:
        raise SystemExit("RAPIDAPI_KEY environment variable is not set.")

    all_jobs = []
    for q in SEARCH_QUERIES:
        try:
            all_jobs.extend(fetch_jobs(q))
        except Exception as e:
            print(f"Warning: query '{q}' failed: {e}")

    all_jobs = dedupe(all_jobs)
    print(f"Fetched {len(all_jobs)} unique jobs.")

    os.makedirs("reports", exist_ok=True)
    today = datetime.date.today().isoformat()
    out_path = f"reports/jobs_{today}.xlsx"
    write_excel(all_jobs, out_path)
    print(f"Saved: {out_path}")

    # JSON feed for the web dashboard (GitHub Pages, /docs folder)
    import json
    os.makedirs("docs", exist_ok=True)
    with open("docs/jobs_data.json", "w", encoding="utf-8") as f:
        json.dump({"generated": today, "jobs": all_jobs}, f, ensure_ascii=False, indent=1)
    print("Saved: docs/jobs_data.json")


if __name__ == "__main__":
    main()
